#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
author: jeff maxey
date: 2018.05.15
purpose: example demonstrating integration of python and alteryx
"""

# // import packages
import os
import subprocess

# // install packages dynamically, adjusting to corporate proxy
def install(name):
    subprocess.call(['pip', '--proxy=cr-proxy.us.aegon.com:9090', 'install', name])

# // install packages from internet as needed
install('pandas')
install('xlrd')
install('openpyxl')

# // then you can import the packages
import pandas as pd
import xlrd
from openpyxl.workbook import Workbook


# // define a function that will list all file paths in specified root directory with specified file extension
def list_file_paths(root_path, ext):
    """
    Purpose: Search all folders in root_path for specific file type.
    :param root_path: The parent directory for traversal search
    :param ext: Specified file extension desired to be searched for.
    :return: A list of files located within root_path that possess specified file extension
    """
    # // create an empty list for storing all file paths
    all_files = []

    # // walk entire directory of root_path; all sub-folders and files will be checked.
    for root, dirs, files in os.walk(root_path):
        for filename in files:
            # // check if file name ends with specified extension
            if filename.lower().endswith(ext):
                # // if it does, generate the full path using os.path.join() and append it to all_files list
                all_files.append(os.path.join(root, filename))

    # // return the list of files
    return all_files


# // specify the directory containing output from alteryx workflow
# // this is a technique that will provide the directory path of where ever this script is located.
dir_path = os.path.dirname(os.path.realpath(__file__))
dir_python_output = os.path.join(dir_path, 'python_output')

# // call the function you created above to collect a list of files in the folder; returns a list
# // returns: ['C:\\Temp\\moses2axis\\docs\\csv\\alteryx_output\\alteryx_output_file_path.csv']
file_list = list_file_paths(root_path=dir_path, ext='.csv')

# // select the file from the list using indexing; returns a string
# // returns: 'C:\\Temp\\moses2axis\\docs\\csv\\alteryx_output\\alteryx_output_file_path.csv'
desired_file = file_list[0]

# // easiest way to read a file is using pandas; you can even do pd.read_excel, pd.read_sql and many more:
df = pd.read_csv(desired_file, sep=',', header=0)


df.columns = [col.replace("?", "") for col in df.columns.tolist()]



def convert_xls_to_xlsx(src_file_path, dst_file_path):
    """
    Purpose: Convert xls file to xlsx. Assumes simple files, no graphics or advanced formatting
    :param src_file_path: 
    :param dst_file_path: 
    :return: 
    """
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index in range(0, len(sheet_names)):
        sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_names[sheet_index]
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row+1, column=col+1).value = sheet_xls.cell_value(row, col)
    book_xlsx.save(dst_file_path)



# // just to show an example of python performing a procedure, lets create an additional column and add the basename
# // of the file path.
# // e.g. 'C:\\Temp\\moses2axis\\docs\\csv\\alteryx_output\\alteryx_output_file_path.csv' --> 'alteryx_output_file_path.csv'
df['file_path_xlsx'] = df['file_path'].apply(lambda x: x.replace('.xls', '.xlsx'))


# // generate a csv file path and output data frame; e.g. 'C:\\Temp\\moses2axis\\moses2axis\\app\\project_directory\\docs\\csv\\xref.csv'
filepathCSV = os.path.join(dir_python_output, "example.csv")
df.to_csv(filepathCSV, sep=",", header=True, index=False)

# // list of xls files
file_list = list_file_paths(dir_path, ext='.xls')

# // for each file in the list of .xls files; call the convert_xls_to_xlsx function.
# // the second parameter creates the desired output file path by replacing ".xls" by ".xlsx"
for file in file_list:
    convert_xls_to_xlsx(file, file.replace(".xls", ".xlsx"))
